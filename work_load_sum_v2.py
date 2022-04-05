#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  work_load_sum_v2.py
#  
#  设计目标：根据提供的班组人员名单、工作计划表，统计各个班组的工作量
#  目前暂时对各级风险进行分类加权。加权系数均为1.  
#  工作流程：
#  1. 读取班组列表banzu.xlsx和计划表jihua.xlsx
#  2. 建立一个相当稀疏的矩阵，存在banzu_index.xlsx里面
#     第一行是日期，第一列是人名，如果某人这一天有工作，则
#     将对应的日期置1
#     
#     设置在两个，包括人名字典和日期字典，存储人名所在行和日期所在列，
#     以便更方便地访问相关的单元格。
#   
#  3. 在打开的班组表中，寻找开始/结束日期、人名、（风险等级暂缺）对应的列
#  4. 处理获取人名单元格并获取列表
#  5. 处理开始单元格和结束单元格，确定需要处理的工作情况表
#  
#  提供：
#  1. banzu.xlsx    班组人员名单
#  2. jihua1.xlsx   汇总的每周工作计划
#
#  上述文件的预处理：
#  1、"jihua1.xlsx"要删除表头，第一行必须是各个标题（可通过title=1修改该设置）
#  2、"jihua1.xlsx"的日期要注意
#  
#  
import string
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import load_workbook
import datetime as dt
import time

import zipfile

# 处理excel中的日期问题
def dateshift(dates):
    delta=dt.timedelta(days=dates)
    today=dt.datetime.strptime('1899-12-30','%Y-%m-%d')+delta
    return dt.datetime.strftime(today,'%Y-%m-%d')


def main(args):
    # Step 1 :
    # 完成基本的配置工作：风险权重，统计时间口径，班组名单，作业计划
	# 风险权重
    xc_risk_weight = {'二级':1,'三级':1,'四级':0.4,'五级':0.2}	
    yq_risk_weight = {'二级':1,'三级':1,'四级':0.2,'五级':0.1}

    # 基准日期
    origin_date = dt.date(dt.date.today().year,1,1)
    
    # 使用从各中心汇总的班组名单
    # 系统、继保、输变电、能源互联网一个科室为一个班组。
    # 物资、人才评价、安全中心整个中心为一个班组
    # 配网分解为直流配网和综合能源两个班组
    wb_banzu = load_workbook(filename = 'banzu.xlsx')
    ws_banzu = wb_banzu['Sheet1']

    
    title = 1
    
    
    # Step2：
    # 根据banzu.xlsx，制作banzu_xc和banxu_yq两个Excel表，
    # 分别统计班组中每个人的现场工作量和园区工作量
    
    
    # Step2.1 找到班组名单banzu中"姓名"那一列
    column_namelist = 1
    activecell = ws_banzu.cell(column = column_namelist,
                                  row = 1)
    while activecell.value != '详细名单':
        column_namelist = column_namelist + 1
        activecell      = ws_banzu.cell(column = column_namelist,
                                           row = 1)                    
                                           
    # Step2.2 找到班组名单banzu中"班组"那一列
    column_banzu = 1
    activecell = ws_banzu.cell(column = column_banzu,
                                  row = 1)    
    while activecell.value != '班组':
        column_banzu = column_banzu + 1
        activecell   = ws_banzu.cell(column = column_banzu,
                                     row = 1)                                   
    # Step2.3 找到班组名单banzu中"人数"那一列
    column_renshu = 1
    activecell = ws_banzu.cell(column = column_renshu,
                                  row = 1)    
    while activecell.value != '人数':
        column_renshu = column_renshu + 1
        activecell   = ws_banzu.cell(column = column_renshu,
                                     row = 1)
    
    # 通过2.1，2.2，2.3，就有了三个变量：
    # column_namelist,column_banzu,column_renshu
    # 实际上也可以直接赋值：
    # column_namelist= 5，column_banzu=3,column_renshu = 6
    # 这三个变量用于后面访问banzu的表格，并更新人员名单。
    
    
    # 新建一个人员列表
    wb = Workbook()
    # 打开工作表中的一个Sheet对象
    ws = wb.active
 
    
    # Step2.4 根据提供的班组名单更新姓名，同时:
    # 初始化班组人员每天的工作量（全部置为0），校核班组人数
    row_p = 2
    activecell = ws_banzu.cell(column = column_namelist,row = row_p)
    banzu_namelist = activecell.value
    name_seq = 2
    
    # 准备空的班组，两个字典，输入班组名称
    # bnazu_renshu[banzu_name.value]表示这个班组有多少人，
    # banzu_zhongxin[banzu_name.value]表示这个班组在哪个中心。
    banzu_renshu             = {}
    banzu_zhongxin           = {}    
 

    # 准备一组空的班组季度工作量，字典备用
    banzu_1jidu_xc = {}
    banzu_2jidu_xc = {}
    banzu_3jidu_xc = {}
    banzu_4jidu_xc = {}
    
    banzu_1jidu_yq = {}
    banzu_2jidu_yq = {}
    banzu_3jidu_yq = {}
    banzu_4jidu_yq = {}
    
        
    
    while banzu_namelist !=None:
		
		# 获取班组名称
        banzu_name = ws_banzu.cell(column = column_banzu,row = row_p)
       
        
        # 2.4.1 先将该班组的工作量置为0
        banzu_1jidu_xc[banzu_name.value] = 0
        banzu_2jidu_xc[banzu_name.value] = 0
        banzu_3jidu_xc[banzu_name.value] = 0
        banzu_4jidu_xc[banzu_name.value] = 0

        banzu_1jidu_yq[banzu_name.value] = 0
        banzu_2jidu_yq[banzu_name.value] = 0
        banzu_3jidu_yq[banzu_name.value] = 0
        banzu_4jidu_yq[banzu_name.value] = 0
        
        
        # 2.4.2 编辑班组人数和班组所在中心的字典，
        # 这两个数据在banzu.xlsx相应的单元格内可以直接访问到。
        banzu_renshu[banzu_name.value] = ws_banzu.cell(
                                    column = column_renshu,
                                    row = row_p).value
        banzu_zhongxin[banzu_name.value] = ws_banzu.cell(
                                    column = 2,
                                    row = row_p).value   
        
        # 2.4.3 对班组名单进行处理，初始化ws这个工作表                                         
        namestring   = str(activecell.value)
        
        # 删除不必要的空白
        namestring.strip()
        
        # 开始处理名单
        name_length = 0
        name_start  = 0
        namelist_index  = 0
        
        # 根据标点来分割人名
        for biaodian in namestring:
            # 判断是否是标点，如果是标点，前面的就是人名。
            # 一般汉字范围是4e00到9fff
            if biaodian < '\u4e00' or biaodian>'\u9fff':
                # 如果是标点，提取一个姓名写入班组表
                if name_length>0:
					
					#写姓名
                    activecell = ws.cell(column = 1,row = name_seq)
                    activecell.value = namestring[name_start:(
                                                name_start+name_length)]
                    
                    #写班组
                    activecell = ws.cell(column = 2,row = name_seq)
                    activecell.value = (ws_banzu.cell(
                                           column = column_banzu,
                                           row    = row_p)).value
                                       
                    #写中心
                    activecell = ws.cell(column = 3,row = name_seq)
                    activecell.value = (ws_banzu.cell(column = 2,
                                           row    = row_p)).value
                    
                    namelist_index = namelist_index + 1
                    name_start  = name_start + name_length + 1
                    name_length = 0
                    name_seq    = name_seq + 1 
                else:
                    name_start  = name_start + 1
                    continue 
            # 如果不是标点        
            else:
                name_length = name_length + 1
        # 处理最后一个人名
        if name_length>0:
            #写姓名
            activecell = ws.cell(column = 1,row = name_seq)
            activecell.value = namestring[name_start:
				                          name_start+name_length]
            
            #写班组信息
            activecell = ws.cell(column = 2,row = name_seq)
            activecell.value = ws_banzu.cell(column = column_banzu,
                                           row    = row_p).value
            #写中心信息
            activecell = ws.cell(column = 3,row = name_seq)
            activecell.value = ws_banzu.cell(column = 2,
                                       row    = row_p).value
            name_start = 0
            name_length = 0
            name_seq = name_seq + 1
            
        
        activecell = ws_banzu.cell(column = column_namelist,
                                      row = row_p + 1)
        row_p = row_p+1
        banzu_namelist = activecell.value
        
    # 根据季度，填写表格中的工作日期，同时开始Step 2.5.1
    # 创建一个字典，以便提供日期后，尽快找到对应的列    
    column_p = 3    
    find_date = {}
    # 一年最多366天
    for dltdate in range(1,367):
        activecell   = ws.cell(column = column_p+dltdate, row = 1)
        delta        = dt.timedelta(days=dltdate)
        activecell.value =dt.datetime.strftime(
            dt.datetime.strptime('2021-12-31',
                                 '%Y-%m-%d')+delta,
                                 '%Y-%m-%d')
        find_date[activecell.value] = column_p+dltdate
    # 创建了1个字典find_date结束,输入一个日期date，
    # 找到对应的列find_date[date]

    # 2.5.2 创建一个字典，以便提供人名后，尽快找到他对应的行
    find_employee = {}
    activecell = ws.cell(column=1,row = 2)
    row_p = 2
    
    while activecell.value !=None:
        find_employee[activecell.value] = row_p
        row_p = row_p + 1
        activecell = ws.cell(column = 1, row = row_p)
        
    wb.save("banzu_xc.xlsx")
    wb.save("banzu_yq.xlsx")
    
    # 打开刚刚生成的两个工作表
    wb_xc = load_workbook("banzu_xc.xlsx")
    ws_xc = wb_xc["Sheet"]
    wb_yq = load_workbook("banzu_yq.xlsx")
    ws_yq = wb_yq["Sheet"] 

    # Step 3 开始处理jihua_jidu这个表格。
    
    # 打开作业计划表    
    wb_jihua = load_workbook(filename = 'jihua_jidu.xlsx')
    ws_jihua_yq = wb_jihua['园区']
    ws_jihua_xc = wb_jihua['现场']
    
    # Step 3.1 获取四个重要参数：
    #  column_employee    描述人员的所在列
    #  column_startdate   描述开始日期的所在列
    #  column_enddate     描述结束日期的所在列
    #  column_risk        描述风险的所在列
    
    
    # 找到班组名单中"姓名"那一列
    # 找到'工作人员'那一列，得到column_employee
    activecell = ws_jihua_yq.cell(column=1,row = title)
    column_employee = 1
    while activecell.value !="工作人员" and activecell.value!=None:
        column_employee   = column_employee + 1
        activecell = ws_jihua_yq.cell(column=column_employee,row = title)
    
    # 找到开始日期那一列，得到column_startdate
    activecell = ws_jihua_yq.cell(column=1,row = title)    
    column_startdate = 1
    while activecell.value !="开始时间" and activecell.value!=None:
        column_startdate  = column_startdate + 1
        activecell = ws_jihua_yq.cell(column=column_startdate,
                                      row = title)   
    
    # 找到结束日期那一列，得到column_enddate
    activecell = ws_jihua_yq.cell(column=1,row = title)    
    column_enddate   = 1
    while activecell.value !="结束时间" and activecell.value!=None:
        column_enddate  = column_enddate + 1
        activecell = ws_jihua_yq.cell(column=column_enddate,row = title)
    
    # 找到风险等级那一列，得到column_risk    
    activecell = ws_jihua_yq.cell(column=1,row = title)    
    column_risk   = 1
    while activecell.value !="风险等级" and activecell.value!=None:
        column_risk  = column_risk + 1
        activecell = ws_jihua_yq.cell(column=column_risk,row = title) 

    
    
    # Step 3.1 对现场工作，从计划的每一行具体内容开始，逐行进行分析
    # 每一行作业对应的是一项作业，因此提取某一项作业，然后对该项作业：
    # （1）是哪些主业人员参与的？
    # （2）参与了哪几天？
    # （3）根据上述情况，在主业人员工作的表格里面，置1（也可以根据风险权重置数）
    
    row_p = title + 1
    activecell = ws_jihua_xc.cell(column=column_employee,row = row_p)
    
    while activecell.value !=None:

        # Step 3.1.1 通过分词手段，在人员单元格中，提取主业员工
        # 找到工作人员名单字符串
        namestring   = str(activecell.value)
       
        # 删除不必要的空白
        namestring.rstrip()
        
        # 提取名字，放入namelist
        namelist       = []
        namelist_index = 0
        name_start     = 0
        name_length    = 0
     
        for biaodian in namestring:
            # 判断是否是标点，如果是标点，前面的就是人名。
            # 一般汉字范围是4e00到9fff
            if biaodian < '\u4e00' or biaodian>'\u9fff':
            # 如果是标点
                if name_length>0:
                    namelist.insert(namelist_index,
                        namestring[name_start:name_start+name_length])
                    # print(namestring[0:name_length-1])
                    namelist_index = namelist_index + 1
                    name_start  = name_start + name_length + 1
                    name_length = 0
                else:
                    name_start = name_start + 1
                    continue 
            # 如果不是标点        
            else:
                name_length = name_length + 1
        
        # 处理开始日期和结束日期
        startdate_cell = ws_jihua_xc.cell(column = column_startdate,
                                       row=row_p)
        enddate_cell   = ws_jihua_xc.cell(column = column_enddate,
                                       row=row_p)
       
        startdate = dt.date.strftime(startdate_cell.value,"%Y-%m-%d")    
        enddate   = dt.date.strftime(enddate_cell.value,"%Y-%m-%d")

        # 在banzu_xc中确定所在的列的范围
        startdate_column = find_date[startdate]
        enddate_column   = find_date[enddate]
        
        # 确定这一项作业的风险等级
        work_risk_cell = ws_jihua_xc.cell(column = column_risk, 
                                       row = row_p)
        work_risk      = work_risk_cell.value
        
                
        # 3.1.2 对于主业员工，统计工作量
        # 遍历namelist中的人的名字，将对应日期置为1
        for employee_name in namelist:
            # 找到对应的人
            if employee_name in find_employee:
                # 找到对应的日期              
                for date_column in range(startdate_column,
                                         enddate_column+1):
                    activecell = ws_xc.cell(column = date_column,
                           row = find_employee[employee_name])
                    
                    # 工作量的统计规则
                    # 
                    #activecell.value = xc_risk_weight[work_risk]
                    
#                    if activecell.value == None:
#                        activecell.value = xc_risk_weight[work_risk]
#                    elif xc_risk_weight[work_risk]>activecell.value:
#                        activecell.value = xc_risk_weight[work_risk]

                    # 备选的统计规则
                    if activecell.value == None:
                        activecell.value = xc_risk_weight[work_risk]
                    elif activecell.value + xc_risk_weight[work_risk]<1:
                        activecell.value = activecell.value + \
                           xc_risk_weight[work_risk]
                    else:
                        activecell.value = 1

        # 完成一个namelist处理后，转到下一行
        row_p      = row_p + 1
        activecell = ws_jihua_xc.cell(column=column_employee,
                                      row = row_p)
    wb_xc.save("banzu_xc.xlsx")
 
 
    # Step 3.2 对园区工作，从计划的每一行具体内容开始，逐行进行分析
    # 每一行作业对应的是一项作业，因此提取某一项作业，然后对该项作业：
    # （1）是哪些主业人员参与的？
    # （2）参与了哪几天？
    # （3）根据上述情况，在主业人员工作的表格里面，置1（也可以根据风险权重置数）
    
    row_p = title + 1        
    activecell = ws_jihua_yq.cell(column=column_employee,row = row_p)
    
    while activecell.value !=None:

        # Step 3.2.1 通过分词手段，在人员单元格中，提取主业员工
        # 找到工作人员名单字符串
        namestring   = str(activecell.value)
       
        # 删除不必要的空白
        namestring.rstrip()
        
        # 提取名字，放入namelist
        namelist       = []
        namelist_index = 0
        name_start     = 0
        name_length    = 0
     
        for biaodian in namestring:
            # 判断是否是标点，如果是标点，前面的就是人名。
            # 一般汉字范围是4e00到9fff
            if biaodian < '\u4e00' or biaodian>'\u9fff':
            # 如果是标点
                if name_length>0:
                    namelist.insert(namelist_index,
                        namestring[name_start:name_start+name_length])
                    # print(namestring[0:name_length-1])
                    namelist_index = namelist_index + 1
                    name_start  = name_start + name_length + 1
                    name_length = 0
                else:
                    name_start = name_start + 1
                    continue 
            # 如果不是标点        
            else:
                name_length = name_length + 1
        
        # 处理开始日期和结束日期
        startdate_cell = ws_jihua_yq.cell(column = column_startdate,
                                       row=row_p)
        enddate_cell   = ws_jihua_yq.cell(column = column_enddate,
                                       row=row_p)
       
        startdate = dt.date.strftime(startdate_cell.value,"%Y-%m-%d")    
        enddate   = dt.date.strftime(enddate_cell.value,"%Y-%m-%d")

        # 在banzu_yq中确定所在的列的范围
        startdate_column = find_date[startdate]
        enddate_column   = find_date[enddate]
        
        # 确定这一项作业的风险等级
        work_risk_cell = ws_jihua_yq.cell(column = column_risk, 
                                       row = row_p)
        work_risk      = work_risk_cell.value
        
                
        # 3.2.2 对于主业员工，统计工作量
        # 遍历namelist中的人的名字，将对应日期置为1
        for employee_name in namelist:
            # 找到对应的人
            if employee_name in find_employee:
                # 找到对应的日期              
                for date_column in range(startdate_column,
                                         enddate_column+1):
                    activecell = ws_yq.cell(column = date_column,
                           row = find_employee[employee_name])
                    
                    # 工作量的统计规则
                    # 
                    #activecell.value = yq_risk_weight[work_risk]
                    
#                    if activecell.value == None:
#                        activecell.value = yq_risk_weight[work_risk]
#                    elif yq_risk_weight[work_risk]>activecell.value:
#                        activecell.value = yq_risk_weight[work_risk]

                    # 备选的统计规则
                    if activecell.value == None:
                        activecell.value = yq_risk_weight[work_risk]
                    elif activecell.value + yq_risk_weight[work_risk]<1:
                        activecell.value = activecell.value + \
                           yq_risk_weight[work_risk]
                    else:
                        activecell.value = 1

        # 完成一个namelist处理后，转到下一行
        row_p      = row_p + 1
        activecell = ws_jihua_yq.cell(column=column_employee,
                                      row = row_p)
    wb_yq.save("banzu_yq.xlsx")
 
    # 4 按列遍历ws工作表，汇总各个班组的工作量    
    # 4.1 平年遍历365天，闰年遍历366天
    if (origin_date.year % 400 == 0) or (
       origin_date.year % 4 ==0 and origin_date.year % 100 != 0):
        rangeday = 366
    else:
        rangeday = 365 

    # 4.2 遍历每一天，统计工作量

    
    # 同时使用ws_xc和ws_yq遍历banzu_xc.xlsx和banzu_yq.xlsx
    for delta_day in range(0,rangeday):
		# 被遍历的那一天
        activeday = origin_date + dt.timedelta(days = delta_day)
        
        # 根据月份判断季度，并将结果汇总到那个季度的字典中
        
        # 4.2.1 统计一季度工作量，月份是1月、2月、3月
        if activeday.month in [1,2,3]: 
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            cell_xc = ws_xc.cell(column = delta_day + 4, row = row_p)
            cell_yq = ws_yq.cell(column = delta_day + 4, row = row_p)
            
            
            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws_xc.cell(column = 1,row = row_p).value != None:
                #如果单元格有数据，则加上
                if cell_xc.value != None:
                    banzu_name = ws_xc.cell(column = 2, 
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_1jidu_xc[banzu_name] = (
                          banzu_1jidu_xc[banzu_name] 
                          + float(str(cell_xc.value)))
                elif cell_yq.value != None:
                    banzu_name = ws_yq.cell(column = 2,
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_1jidu_yq[banzu_name] = (
                          banzu_1jidu_yq[banzu_name] 
                          + float(str(cell_yq.value)))
                    
                # 循环控制，到下一行          
                row_p = row_p + 1
                cell_xc = ws_xc.cell(column = delta_day + 4, 
                                     row = row_p)
                cell_yq = ws_yq.cell(column = delta_day + 4, 
                                     row = row_p)
        # 二季度
        elif activeday.month in [4,5,6]:
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            cell_xc = ws_xc.cell(column = delta_day + 4, row = row_p)
            cell_yq = ws_yq.cell(column = delta_day + 4, row = row_p)
            
            
            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws_xc.cell(column = 1,row = row_p).value != None:
                #如果单元格有数据，则加上
                if cell_xc.value != None:
                    banzu_name = ws_xc.cell(column = 2, 
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_2jidu_xc[banzu_name] = (
                          banzu_2jidu_xc[banzu_name] 
                          + float(str(cell_xc.value)))
                elif cell_yq.value != None:
                    banzu_name = ws_yq.cell(column = 2,
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_2jidu_yq[banzu_name] = (
                          banzu_2jidu_yq[banzu_name] 
                          + float(str(cell_yq.value)))
                    
                # 循环控制，到下一行          
                row_p = row_p + 1
                cell_xc = ws_xc.cell(column = delta_day + 4, 
                                     row = row_p)
                cell_yq = ws_yq.cell(column = delta_day + 4, 
                                     row = row_p)
        # 三季度                             
        elif activeday.month in [7,8,9]:
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            cell_xc = ws_xc.cell(column = delta_day + 4, row = row_p)
            cell_yq = ws_yq.cell(column = delta_day + 4, row = row_p)
            
            
            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws_xc.cell(column = 1,row = row_p).value != None:
                #如果单元格有数据，则加上
                if cell_xc.value != None:
                    banzu_name = ws_xc.cell(column = 2, 
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_3jidu_xc[banzu_name] = (
                          banzu_3jidu_xc[banzu_name] 
                          + float(str(cell_xc.value)))
                elif cell_yq.value != None:
                    banzu_name = ws_yq.cell(column = 2,
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_3jidu_yq[banzu_name] = (
                          banzu_3jidu_yq[banzu_name] 
                          + float(str(cell_yq.value)))
                    
                # 循环控制，到下一行          
                row_p = row_p + 1
                cell_xc = ws_xc.cell(column = delta_day + 4, 
                                     row = row_p)
                cell_yq = ws_yq.cell(column = delta_day + 4, 
                                     row = row_p)
        # 四季度                             
        elif activeday.month in [10,11,12]:
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            cell_xc = ws_xc.cell(column = delta_day + 4, row = row_p)
            cell_yq = ws_yq.cell(column = delta_day + 4, row = row_p)
            
            
            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws_xc.cell(column = 1,row = row_p).value != None:
                #如果单元格有数据，则加上
                if cell_xc.value != None:
                    banzu_name = ws_xc.cell(column = 2, 
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_4jidu_xc[banzu_name] = (
                          banzu_4jidu_xc[banzu_name] 
                          + float(str(cell_xc.value)))
                elif cell_yq.value != None:
                    banzu_name = ws_yq.cell(column = 2,
                                            row = row_p).value
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_4jidu_yq[banzu_name] = (
                          banzu_4jidu_yq[banzu_name] 
                          + float(str(cell_yq.value)))
                    
                # 循环控制，到下一行          
                row_p = row_p + 1
                cell_xc = ws_xc.cell(column = delta_day + 4, 
                                     row = row_p)
                cell_yq = ws_yq.cell(column = delta_day + 4, 
                                     row = row_p)



#    print(banzu_3jidu_xc,banzu_3jidu_yq)
    
    
    # 4.3 编写汇总表
    #整理下格式
    wb_huizong = Workbook()
    ws_1jidu = wb_huizong.create_sheet("一季度",0)
    ws_2jidu = wb_huizong.create_sheet("二季度",1)    
    ws_3jidu = wb_huizong.create_sheet("三季度",2)    
    ws_4jidu = wb_huizong.create_sheet("四季度",3) 
    
    for ws_title in wb_huizong:     
        ws_title['A1']='序号'
        ws_title['B1']='班组'
        ws_title['C1']='部门/中心'
        ws_title['D1']='班组人数'
        ws_title['E1']='全口径总工作量'
        ws_title['F1']='全口径平均工作量'
        ws_title['G1']='现场工作总量'
        ws_title['H1']='现场人均工作量'
        ws_title['I1']='园区实验室工作总量'
        ws_title['J1']='园区实验室人均工作量'
    
    seq_No = 1
    for banzu in banzu_1jidu_xc:
        ws_1jidu.cell(column=1,row = seq_No + 1).value = seq_No		
        # 填写班组名
        ws_1jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_1jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写班组人数            
        ws_1jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_renshu[banzu]

        # 填写总工作量
        ws_1jidu.cell(column=5,row = seq_No + 1
         ).value = banzu_1jidu_xc[banzu]+banzu_1jidu_yq[banzu]
        
        # 计算平均工作量           
        ws_1jidu.cell(column=6,row = seq_No + 1
             ).value = (banzu_1jidu_xc[banzu]+banzu_1jidu_yq[banzu
                           ])/banzu_renshu[banzu]
                           
        # 填写现场工作量
        ws_1jidu.cell(column=7,row = seq_No + 1
         ).value = banzu_1jidu_xc[banzu]
        
        # 计算现场平均工作量           
        ws_1jidu.cell(column=8,row = seq_No + 1
             ).value = banzu_1jidu_xc[banzu
                           ]/banzu_renshu[banzu]
                           
        # 填写园区总工作量
        ws_1jidu.cell(column=9,row = seq_No + 1
         ).value = banzu_1jidu_yq[banzu]
        
        # 计算园区平均工作量           
        ws_1jidu.cell(column=10,row = seq_No + 1
             ).value = banzu_1jidu_yq[banzu
                           ]/banzu_renshu[banzu]
                   
                           
        seq_No = seq_No + 1
        
        
    seq_No = 1
    for banzu in banzu_2jidu_xc:
        ws_2jidu.cell(column=1,row = seq_No + 1).value = seq_No
        
        # 填写班组名
        ws_2jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_2jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写班组人数            
        ws_2jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_renshu[banzu]

        # 填写总工作量
        ws_2jidu.cell(column=5,row = seq_No + 1
         ).value = banzu_2jidu_xc[banzu]+banzu_2jidu_yq[banzu]
        
        # 计算平均工作量           
        ws_2jidu.cell(column=6,row = seq_No + 1
             ).value = (banzu_2jidu_xc[banzu]+banzu_2jidu_yq[banzu
                           ])/banzu_renshu[banzu]
                           
        # 填写现场工作量
        ws_2jidu.cell(column=7,row = seq_No + 1
         ).value = banzu_2jidu_xc[banzu]
        
        # 计算现场平均工作量           
        ws_2jidu.cell(column=8,row = seq_No + 1
             ).value = banzu_2jidu_xc[banzu
                           ]/banzu_renshu[banzu]
                           
        # 填写园区总工作量
        ws_2jidu.cell(column=9,row = seq_No + 1
         ).value = banzu_2jidu_yq[banzu]
        
        # 计算园区平均工作量           
        ws_2jidu.cell(column=10,row = seq_No + 1
             ).value = banzu_2jidu_yq[banzu
                           ]/banzu_renshu[banzu]
                   
                           
        seq_No = seq_No + 1
    
    #填写三季度
    seq_No = 1
    for banzu in banzu_3jidu_xc:
        ws_3jidu.cell(column=1,row = seq_No + 1).value = seq_No
        # 填写班组名
        ws_3jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_3jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写班组人数            
        ws_3jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_renshu[banzu]

        # 填写总工作量
        ws_3jidu.cell(column=5,row = seq_No + 1
         ).value = banzu_3jidu_xc[banzu]+banzu_3jidu_yq[banzu]
        
        # 计算平均工作量           
        ws_3jidu.cell(column=6,row = seq_No + 1
             ).value = (banzu_3jidu_xc[banzu]+banzu_3jidu_yq[banzu
                           ])/banzu_renshu[banzu]
                           
        # 填写现场工作量
        ws_3jidu.cell(column=7,row = seq_No + 1
         ).value = banzu_3jidu_xc[banzu]
        
        # 计算现场平均工作量           
        ws_3jidu.cell(column=8,row = seq_No + 1
             ).value = banzu_3jidu_xc[banzu
                           ]/banzu_renshu[banzu]
                           
        # 填写园区总工作量
        ws_3jidu.cell(column=9,row = seq_No + 1
         ).value = banzu_3jidu_yq[banzu]
        
        # 计算园区平均工作量           
        ws_3jidu.cell(column=10,row = seq_No + 1
             ).value = banzu_3jidu_yq[banzu
                           ]/banzu_renshu[banzu]
        seq_No = seq_No + 1           

    #填写四季度
    seq_No = 1        
    for banzu in banzu_4jidu_xc:
        ws_4jidu.cell(column=1,row = seq_No + 1).value = seq_No
        # 填写班组名
        ws_4jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_4jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写班组人数            
        ws_4jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_renshu[banzu]

        # 填写总工作量
        ws_4jidu.cell(column=5,row = seq_No + 1
         ).value = banzu_4jidu_xc[banzu]+banzu_4jidu_yq[banzu]
        
        # 计算平均工作量           
        ws_4jidu.cell(column=6,row = seq_No + 1
             ).value = (banzu_4jidu_xc[banzu]+banzu_4jidu_yq[banzu
                           ])/banzu_renshu[banzu]
                           
        # 填写现场工作量
        ws_4jidu.cell(column=7,row = seq_No + 1
         ).value = banzu_4jidu_xc[banzu]
        
        # 计算现场平均工作量           
        ws_4jidu.cell(column=8,row = seq_No + 1
             ).value = banzu_4jidu_xc[banzu
                           ]/banzu_renshu[banzu]
                           
        # 填写园区总工作量
        ws_4jidu.cell(column=9,row = seq_No + 1
         ).value = banzu_4jidu_yq[banzu]
        
        # 计算园区平均工作量           
        ws_4jidu.cell(column=10,row = seq_No + 1
             ).value = banzu_4jidu_yq[banzu
                           ]/banzu_renshu[banzu]
                   
                           
                           
        seq_No = seq_No + 1

    
    wb_huizong.save("汇总表.xlsx")   
    
    
    return 0

if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv))

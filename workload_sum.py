#!/usr/bin/python3
# -*- coding: UTF-8 -*-
#
#  safetyaward.py
#  
#  设计目标：根据提供的班组人员名单、工作计划表，统计各个班组的工作量
#  目前暂时对各级风险进行分类加权。加权系数均为1.  
#  
#  工作流程：
#  1. 读取班组列表和计划表
#  2. 设置在班组表中的字典，包括人名字典和日期字典，
#     以便更方便地访问相关的单元格。
#  3. 在打开的班组表中，寻找开始/结束日期、人名、（风险等级暂缺）对应的列
#  4. 处理获取人名单元格并获取列表
#  5. 处理开始单元格和结束单元格，确定需要处理的工作情况表
#  
#  提供：
#  1. banzu.xlsx    班组名单
#  2. jihua1.xlsx   汇总的每周工作计划
#  （风险权重）为升级功能
#  输出结果：
#  汇总表.xlsx 包含各个季度的情况，仅有当季度的是有效数据
#
#  中间过程：
#  1. banzu_index.xlsx   精确到每个人的工作量统计
#  2. 字典和列表：
#         banzu：       列表，存储班组名称
#         banzu_renshu：字典，根据班组名，找到对应班组的人数
#         banzu_zhongxin： 根据班组找到对应中心
#         汇总的四个季度的工作量
#         banzu_1jidu_gongzuoliang:
#         banzu_2jidu_gongzuoliang:
#         banzu_3jidu_gongzuoliang:
#         banzu_4jidu_gongzuoliang:
#         
#         find_employee: 根据姓名找到在banzu_index.xlsx的行
#         find_date：    根据提供的日期找到在banzu_index.xlsx的列
#         
#  3. wb和ws对象：
#  
#

import string
from openpyxl import workbook
from openpyxl import load_workbook
import datetime
import time

import zipfile

# 处理excel中的日期问题
def date(dates):
    delta=datetime.timedelta(days=dates)
    today=datetime.datetime.strptime('1899-12-30','%Y-%m-%d')+delta
    return datetime.datetime.strftime(today,'%Y-%m-%d')

def main(args):

    
	# 风险权重
    risk_weight = {'一级':1,'二级':1,'三级':1,'四级':1}	

    # 基础日期
    origin_date = datetime.date(datetime.date.today().year,1,1)
    
    # 使用从各中心汇总的班组名单
    wb_banzu = load_workbook(filename = 'banzu.xlsx')
    ws_banzu = wb_banzu['Sheet1']
    
    # 找到班组名单中"姓名"那一列
    column_namelist = 1
    activecell = ws_banzu.cell(column = column_namelist,
                                  row = 1)
    while activecell.value != '详细名单':
        column_namelist = column_namelist + 1
        activecell      = ws_banzu.cell(column = column_namelist,
                                           row = 1)                    
                                           
    # 找到"班组"那一列
    column_banzu = 1
    activecell = ws_banzu.cell(column = column_banzu,
                                  row = 1)    
    while activecell.value != '班组':
        column_banzu = column_banzu + 1
        activecell   = ws_banzu.cell(column = column_banzu,
                                     row = 1)                                   
    
    # 打开班组列表
    wb = load_workbook(filename = 'banzu_index.xlsx')
    # 打开工作表中的一个Sheet对象
    ws = wb['Sheet1']
 
    # 找到"人数"那一列
    column_renshu = 1
    activecell = ws_banzu.cell(column = column_renshu,
                                  row = 1)    
    while activecell.value != '人数':
        column_renshu = column_renshu + 1
        activecell   = ws_banzu.cell(column = column_renshu,
                                     row = 1)
    
    # 根据提供的班组名单更新姓名，同时初始化班组工作量，校核班组人数
    row_p = 2
    activecell = ws_banzu.cell(column = column_namelist,row = row_p)
    banzu_namelist = activecell.value
    name_seq = 2
 
    # 准备一组空的班组季度工作量字典
    banzu_1jidu_gongzuoliang = {}
    banzu_2jidu_gongzuoliang = {}
    banzu_3jidu_gongzuoliang = {}
    banzu_4jidu_gongzuoliang = {}
    
    # 准备空的班组
    banzu_renshu             = {}
    banzu_zhongxin           = {}    
    while banzu_namelist !=None:
		
        banzu_name = ws_banzu.cell(column = column_banzu,row = row_p)
        
        banzu_1jidu_gongzuoliang[banzu_name.value] = 0
        banzu_2jidu_gongzuoliang[banzu_name.value] = 0
        banzu_3jidu_gongzuoliang[banzu_name.value] = 0
        banzu_4jidu_gongzuoliang[banzu_name.value] = 0
        
        banzu_renshu[banzu_name.value] = ws_banzu.cell(
                                              column = column_renshu,
                                              row = row_p).value
        banzu_zhongxin[banzu_name.value] = ws_banzu.cell(
                                                 column = 2,
                                              row = row_p).value   
                                                 
        namestring   = str(activecell.value)
        
        # 删除不必要的空白
        namestring.strip()
        
        # 开始处理名单
        name_length = 0
        name_start  = 0
        namelist_index  = 0
        
        for biaodian in namestring:
            # 判断是否是标点，如果是标点，前面的就是人名。
            # 一般汉字范围是4e00到9fff
            if biaodian < '\u4e00' or biaodian>'\u9fff':
                # 如果是标点，提取一个姓名写入班组表
                if name_length>0:
					
					#写姓名
                    activecell = ws.cell(column = 1,row = name_seq)
                    activecell.value = namestring[name_start:(
                                                name_start+name_length)]
                    
                    #写班组
                    activecell = ws.cell(column = 2,row = name_seq)
                    activecell.value = (ws_banzu.cell(
                                           column = column_banzu,
                                           row    = row_p)).value
                                       
                    #写中心
                    activecell = ws.cell(column = 3,row = name_seq)
                    activecell.value = (ws_banzu.cell(column = 2,
                                           row    = row_p)).value
                    
                    namelist_index = namelist_index + 1
                    name_start  = name_start + name_length + 1
                    name_length = 0
                    name_seq    = name_seq + 1 
                else:
                    name_start  = name_start + 1
                    continue 
            # 如果不是标点        
            else:
                name_length = name_length + 1
        # 处理最后一个人名
        if name_length>0:
            #写姓名
            activecell = ws.cell(column = 1,row = name_seq)
            activecell.value = namestring[name_start:
				                          name_start+name_length]
            
            #写班组信息
            activecell = ws.cell(column = 2,row = name_seq)
            activecell.value = ws_banzu.cell(column = column_banzu,
                                           row    = row_p).value
            #写中心信息
            activecell = ws.cell(column = 3,row = name_seq)
            activecell.value = ws_banzu.cell(column = 2,
                                       row    = row_p).value
            name_start = 0
            name_length = 0
            name_seq = name_seq + 1
            
        
        activecell = ws_banzu.cell(column = column_namelist,
                                      row = row_p + 1)
        row_p = row_p+1
        banzu_namelist = activecell.value

    # 创建一个字典，以便提供人名后，尽快找到他对应的行
    find_employee = {}
    activecell = ws.cell(column=1,row = 2)
    row_p = 2
    
    while activecell.value !=None:
        find_employee[activecell.value] = row_p
        row_p = row_p + 1
        activecell = ws.cell(column = 1, row = row_p)
    

    # 根据季度，填写表格中的工作日期
    # 同时创建一个字典，以便提供日期后，尽快找到对应的列    
    column_p = 3    
    find_date = {}
    # 一年最多366天
    for dltdate in range(1,367):
        activecell   = ws.cell(column = column_p+dltdate, row = 1)
        delta        = datetime.timedelta(days=dltdate)
        activecell.value =datetime.datetime.strftime(
            datetime.datetime.strptime('2019-12-31','%Y-%m-%d')+delta,
            '%Y-%m-%d')
        find_date[activecell.value] = column_p+dltdate
    # 创建2个字典结束
    
    # 打开作业计划表
    
    wb_jihua = load_workbook(filename = 'jihua1.xlsx')
    ws_jihua = wb_jihua['现场']
    
    title = 2
    
    # 找到'工作人员'那一列
    activecell = ws_jihua.cell(column=1,row = title)
    column_employee = 1
    while activecell.value !="工作人员" and activecell.value!=None:
        column_employee   = column_employee + 1
        activecell = ws_jihua.cell(column=column_employee,row = title)
    
    # 找到开始日期那一列
    activecell = ws_jihua.cell(column=1,row = title)    
    column_startdate = 1
    while activecell.value !="开始时间" and activecell.value!=None:
        column_startdate  = column_startdate + 1
        activecell = ws_jihua.cell(column=column_startdate,row = title)
    
    # 找到结束日期那一列
    activecell = ws_jihua.cell(column=1,row = title)    
    column_enddate   = 1
    while activecell.value !="结束时间" and activecell.value!=None:
        column_enddate  = column_enddate + 1
        activecell = ws_jihua.cell(column=column_enddate,row = title)
    
    # 找到风险等级那一列    
    activecell = ws_jihua.cell(column=1,row = title)    
    column_risk   = 1
    while activecell.value !="风险等级" and activecell.value!=None:
        column_risk  = column_risk + 1
        activecell = ws_jihua.cell(column=column_risk,row = title) 

    # 从具体内容行开始，逐行进行分析
    row_p = title + 1        
    activecell = ws_jihua.cell(column=column_employee,row = row_p)
    
    while activecell.value !=None:
        # 分词手段找到人
        # 找到工作人员名单字符串
        namestring   = str(activecell.value)
       
        # 删除不必要的空白
        namestring.rstrip()
        
        # 提取名字，放入namelist
        namelist       = []
        namelist_index = 0
        name_start     = 0
        name_length    = 0
     
        for biaodian in namestring:
            # 判断是否是标点，如果是标点，前面的就是人名。
            # 一般汉字范围是4e00到9fff
            if biaodian < '\u4e00' or biaodian>'\u9fff':
            # 如果是标点
                if name_length>0:
                    namelist.insert(namelist_index,
                        namestring[name_start:name_start+name_length])
                    # print(namestring[0:name_length-1])
                    namelist_index = namelist_index + 1
                    name_start  = name_start + name_length + 1
                    name_length = 0
                else:
                    name_start = name_start + 1
                    continue 
            # 如果不是标点        
            else:
                name_length = name_length + 1
        
        # 处理开始日期和结束日期
        startdate_cell = ws_jihua.cell(column = column_startdate,
                                       row=row_p)
        enddate_cell   = ws_jihua.cell(column = column_enddate,
                                       row=row_p)
                                       
        startdate = date(startdate_cell.value)
        enddate   = date(enddate_cell.value)  
        
        startdate_column = find_date[startdate]
        enddate_column   = find_date[enddate]
        # 确定这一项作业的风险等级
        work_risk_cell = ws_jihua.cell(column = column_risk, 
                                       row = row_p)
        work_risk      = work_risk_cell.value
                
        # 对于主业员工，统计工作量
        # 遍历namelist中的人的名字，将对应日期置为1
        for employee_name in namelist:
            if employee_name in find_employee:

                for date_column in range(startdate_column,
                                         enddate_column+1):
                    activecell = ws.cell(column = date_column,
                           row = find_employee[employee_name])
                           
                    activecell.value = risk_weight[work_risk]
        # 完成一个namelist处理后，转到下一行
        row_p      = row_p + 1
        activecell = ws_jihua.cell(column=column_employee,row = row_p)
    
    # 保存工作表
    wb.save('banzu_index.xlsx')
    
    # 按列遍历ws工作表，汇总各个班组的工作量    
    # 平年遍历365天，闰年遍历366天
    if (origin_date.year % 400 == 0) or (
       origin_date.year % 4 ==0 and origin_date.year % 100 != 0):
        rangeday = 366
    else:
        rangeday = 365 
    
    # 遍历每一天，统计
    for delta_day in range(0,rangeday):
		# 被遍历的那一天
        activeday = origin_date + datetime.timedelta(days = delta_day)
        
        # 根据月份判断季度，并将结果汇总到那个季度的字典中
        # 一季度
        if activeday.month in [1,2,3]: 
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            activecell = ws.cell(column = delta_day + 4, row = row_p)
            
            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws.cell(column = 1,row = row_p).value != None:
				#如果单元格有数据，则加上
                if activecell.value != None:
                    # 有工作量可被计算的班组名称
                    banzu_name = ws.cell(column = 2, row = row_p).value
                    
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_1jidu_gongzuoliang[banzu_name] = (
                          banzu_1jidu_gongzuoliang[banzu_name] 
                          + int(str(activecell.value)))
                # 循环控制，到下一行          
                row_p = row_p + 1
                activecell = ws.cell(column = delta_day + 4, 
                                     row = row_p)                         
        # 二季度
        elif activeday.month in [4,5,6]:
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            activecell = ws.cell(column = delta_day + 4, row = row_p)

            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws.cell(column = 1,row = row_p).value != None:
				#如果单元格有数据，则加上
                if activecell.value != None:
                    # 有工作量可被计算的班组名称
                    banzu_name = ws.cell(column = 2, row = row_p).value
                
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_2jidu_gongzuoliang[banzu_name] = (
                       banzu_2jidu_gongzuoliang[banzu_name] 
                       + int(str(activecell.value)))

                # 循环控制，到下一行          
                row_p = row_p + 1
                activecell = ws.cell(column = delta_day + 4, 
                                     row = row_p)
        # 三季度                             
        elif activeday.month in [7,8,9]:
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            activecell = ws.cell(column = delta_day + 4, row = row_p)

            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws.cell(column = 1,row = row_p).value != None:

				#如果单元格有数据，则加上
                if activecell.value != None:
                    # 有工作量可被计算的班组名称
                    banzu_name = ws.cell(column = 2, row = row_p).value
                
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_3jidu_gongzuoliang[banzu_name] = (
                       banzu_3jidu_gongzuoliang[banzu_name] 
                       + int(str(activecell.value)))
                # 循环控制，到下一行          
                row_p = row_p + 1
                activecell = ws.cell(column = delta_day + 4, 
                                     row = row_p)
        # 四季度                             
        elif activeday.month in [10,11,12]:
            # 每一行对应了一个人和一个班组计数，从第二行开始统计
            # +4 是因为前三列分别为人名、班组名和中心名，第四列开始才是第一个日期
            row_p = 2
            activecell = ws.cell(column = delta_day + 4, row = row_p)

            # 从上到下逐行累加，直到统计完所有人，按人名来判断循环结束
            while ws.cell(column = 1,row = row_p).value != None:

				#如果单元格有数据，则加上
                if activecell.value != None:
                    # 有工作量可被计算的班组名称
                    banzu_name = ws.cell(column = 2, row = row_p).value
                
                    # 每个班组的工作量保存在一个字典里，
                    # 通过名称更新字典中的值
                    banzu_4jidu_gongzuoliang[banzu_name] = (
                       banzu_4jidu_gongzuoliang[banzu_name] 
                       + int(str(activecell.value)))

                # 循环控制，到下一行          
                row_p = row_p + 1
                activecell = ws.cell(column = delta_day + 4, 
                                     row = row_p)
    # 将工作量写入汇总工作表
    # 将统计结果写入汇总表
    wb_huizong = load_workbook('汇总.xlsx')
    ws_1jidu = wb_huizong['一季度']
    # 表格项目
    ws_1jidu['A1']='序号'
    ws_1jidu['B1']='班组'
    ws_1jidu['C1']='部门/中心'
    ws_1jidu['D1']='总工作量'
    ws_1jidu['E1']='班组人数'
    ws_1jidu['F1']='平均工作量'
    ws_1jidu['G1']='初评结果'
    
    seq_No = 1
    for banzu in banzu_1jidu_gongzuoliang:
        # 填写班组名
        ws_1jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_1jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写总工作量
        ws_1jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_1jidu_gongzuoliang[banzu]
        
        
        # 填写班组人数            
        ws_1jidu.cell(column=5,row = seq_No + 1
                    ).value = banzu_renshu[banzu]
        
        # 计算平均工作量           
        ws_1jidu.cell(column=6,row = seq_No + 1
             ).value = banzu_1jidu_gongzuoliang[
                                  banzu]/banzu_renshu[banzu]
        seq_No = seq_No + 1
   
    ws_2jidu = wb_huizong['二季度']
    # 表格项目
    ws_2jidu['A1']='序号'
    ws_2jidu['B1']='班组'
    ws_2jidu['C1']='部门/中心'
    ws_2jidu['D1']='总工作量'
    ws_2jidu['E1']='班组人数'
    ws_2jidu['F1']='平均工作量'
    ws_2jidu['G1']='初评结果'
    
    seq_No = 1
    for banzu in banzu_2jidu_gongzuoliang:
        # 填写班组名
        ws_2jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_2jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写总工作量
        ws_2jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_2jidu_gongzuoliang[banzu]
        
        
        # 填写班组人数            
        ws_2jidu.cell(column=5,row = seq_No + 1
                    ).value = banzu_renshu[banzu]
        
        # 计算平均工作量           
        ws_2jidu.cell(column=6,row = seq_No + 1
             ).value = banzu_2jidu_gongzuoliang[
                                  banzu]/banzu_renshu[banzu]
        seq_No = seq_No + 1
   
    ws_3jidu = wb_huizong['三季度']
    
    # 表格项目
    ws_3jidu['A1']='序号'
    ws_3jidu['B1']='班组'
    ws_3jidu['C1']='部门/中心'
    ws_3jidu['D1']='总工作量'
    ws_3jidu['E1']='班组人数'
    ws_3jidu['F1']='平均工作量'
    ws_3jidu['G1']='初评结果'
    
    seq_No = 1
    for banzu in banzu_3jidu_gongzuoliang:
        # 填写班组名
        ws_3jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_3jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写总工作量
        ws_3jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_3jidu_gongzuoliang[banzu]
        
        
        # 填写班组人数            
        ws_3jidu.cell(column=5,row = seq_No + 1
                    ).value = banzu_renshu[banzu]
        
        # 计算平均工作量           
        ws_3jidu.cell(column=6,row = seq_No + 1
             ).value = banzu_3jidu_gongzuoliang[
                                  banzu]/banzu_renshu[banzu]
        seq_No = seq_No + 1
        
    ws_4jidu = wb_huizong['四季度']
    
    # 表格项目
    ws_4jidu['A1']='序号'
    ws_4jidu['B1']='班组'
    ws_4jidu['C1']='部门/中心'
    ws_4jidu['D1']='总工作量'
    ws_4jidu['E1']='班组人数'
    ws_4jidu['F1']='平均工作量'
    ws_4jidu['G1']='初评结果'
    
    seq_No = 1
    for banzu in banzu_4jidu_gongzuoliang:
        # 填写班组名
        ws_4jidu.cell(column=2,row = seq_No + 1).value = banzu

        # 填写所在中心
        ws_4jidu.cell(column=3,row = seq_No + 1
                    ).value = banzu_zhongxin[banzu]
        
        # 填写总工作量
        ws_4jidu.cell(column=4,row = seq_No + 1
                    ).value = banzu_4jidu_gongzuoliang[banzu]
        
        # 填写班组人数
        ws_4jidu.cell(column=5,row = seq_No + 1
                    ).value = banzu_renshu[banzu]
        
        # 计算平均工作量           
        ws_4jidu.cell(column=6,row = seq_No + 1
             ).value = banzu_4jidu_gongzuoliang[
                                  banzu]/banzu_renshu[banzu]
        seq_No = seq_No + 1
       
    wb_huizong.save('汇总表.xlsx')
    return 0

if __name__ == '__main__':
    import sys
    sys.exit(main(sys.argv))
